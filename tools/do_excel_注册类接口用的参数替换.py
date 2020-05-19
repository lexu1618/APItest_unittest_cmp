import pandas as pd
from openpyxl import load_workbook
from config.project_path import *
from tools.get_global_data import GetData
from tools.do_regx import DoRegx

class DoExcel:
    @classmethod
    def get_data(cls,filename,sheetname):
        wb =load_workbook(filename)
        # mode = eval(ReadConfig.get_config(case_config_path,"MODE","mode"))
        sheet = wb[sheetname]  # 表单名
        test_data =[]

        tel = getattr(GetData,"NoRegTel")

        #利用python查询数据库的方式拿到最大的手机号---这里可以加，也可以放到getdata里
        #
        #
        #

        for i in range(2, sheet.max_row + 1):
            row_data = {}  # 字典
            row_data["case_id"] = sheet.cell(i, 1).value
            row_data["module"] = sheet.cell(i, 2).value
            row_data["title"] = sheet.cell(i, 3).value
            row_data["url"] = sheet.cell(i, 4).value
            # row_data["data"] = sheet.cell(i,5).value


            #字符串替换（用于excel中替换一些固定的字符串，例如管理员账号密码等固定数据）    find 函数找不到目标字符串时返回为-1，能找到返回字符串的索引
            if sheet.cell(i,5).value.find("${tel_1}")!= -1:    #if h后面非空  非零  成立的表达式  都为True，只要是True，if下面的代码都会执行
                row_data["data"] = sheet.cell(i,5).value.replace("${tel_1}",str(tel))
                tel += 1

            # elif sheet.cell(i,5).value.find("${admin_tel}")!= -1:
            #     row_data["data"] = sheet.cell(i,5).value.replace("${tel}",str(tel+1))
            #
            # elif sheet.cell(i,5).value.find("${abcdefg}") != -1:
            #     row_data["data"] = sheet.cell(i, 5).value.replace("${tel}", str(getattr(GetData,"XXXXXX")))
            #
            #
            # elif sheet.cell(i, 5).value.find("${xxxxx}") != -1:
            #
            #     row_data["data"] = sheet.cell(i, 5).value.replace("${tel}", str(getattr(GetData, "XXXXXX")))
            #
            # else:
            #     row_data["data"] = sheet.cell(i,5).value

            #改为正则   一行替代多行
            else:
                row_data["data"] = DoRegx.do_regx(sheet.cell(i, 5).value)


            row_data["http_method"] = sheet.cell(i,6).value
            row_data["expect"] = sheet.cell(i, 7).value
            row_data["sheet_name"] = sheetname
            test_data.append(row_data)
            #更新手机号   针对EXCEL 手机号码更新操作
            cls.update_tel(tel+2,filename,"init")


            # else:
            #     row_data = {}  # 字典
            #     row_data["case_id"] = sheet.cell(case_id+1, 1).value
            #     row_data["username"] = sheet.cell(case_id+1, 2).value
            #     row_data["password"] = sheet.cell(case_id+1, 3).value
            #     row_data["email"] = sheet.cell(i, 4).value
            #     row_data["sheet_name"] = sheetname
            #     test_data.append(row_data)

        return test_data

    @staticmethod
    def write_back(filename,sheetname,i,result,TestResult):#回写数据
        wb = load_workbook(filename)
        sheet = wb[sheetname]
        sheet.cell(i+1,8).value = result
        sheet.cell(i + 1,9).value = TestResult
        wb.save(filename)#保存结果

    #更新EXCEL   init表里的未注册手机号的数据
    @classmethod
    def update_tel(cls,tel,filename,sheet_name):
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        sheet.cell(2,1).value = tel
        wb.save(filename)


    #获取未注册的手机号
    # def get_tel(self):
    #     pass



if __name__ == '__main__':
    pass