import pandas as pd
from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools.do_regx import DoRegx
from tools.my_log import MyLog
from collections import namedtuple

my_logger = MyLog()

#openpyxl操作excel时，行号和列号都是从1开始计算的
class DoExcel:

    def __init__(self,filename,sheet_name):
        try:
            self.filename = filename
            self.sheet_name = sheet_name
            self.wb = load_workbook(self.filename)
            if self.sheet_name is None:
                self.work_sheet = self.wb.active
            else:
                self.work_sheet = self.wb[self.sheet_name]
        except FileNotFoundError as e:
            my_logger.error("文件不存在")
            raise e

    def get_max_row_num(self):
        """获取最大行号"""
        max_row_num = self.work_sheet.max_row
        return max_row_num

    def get_max_column_num(self):
        """获取最大列号"""
        max_column = self.work_sheet.max_column
        return max_column

    def get_cell_value(self, coordinate=None, row=None, column=None):
        """获取指定单元格的数据    coordinate为单元格的位置，例如A1  B3 等"""
        if coordinate is not None:
            try:
                return self.work_sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and row is not None and column is not None:
            if isinstance(row, int) and isinstance(column, int):
                return self.work_sheet.cell(row=row, column=column).value
            else:
                raise TypeError('row and column must be type int')
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def get_row_value(self, row):
        """获取某一行的数据"""
        column_num = self.get_max_column_num()
        row_value = []
        if isinstance(row, int):
            for column in range(1, column_num + 1):
                values_row = self.work_sheet.cell(row, column).value
                row_value.append(values_row)
            return row_value
        else:
            raise TypeError('row must be type int')

    def get_column_value(self, column):
        """获取某一列数据"""
        row_num = self.get_max_column_num()
        column_value = []
        if isinstance(column, int):
            for row in range(1, row_num + 1):
                values_column = self.work_sheet.cell(row, column).value
                column_value.append(values_column)
            return column_value
        else:
            raise TypeError('column must be type int')

    # def get_all_value_1(self):
    #     """获取指定表单的所有数据(除去表头)"""
    #     max_row_num = self.get_max_row_num()
    #     max_column = self.get_max_column_num()
    #     values = []
    #     for row in range(2, max_row_num + 1):
    #         value_list = []
    #         for column in range(1, max_column + 1):
    #             value = self.work_sheet.cell(row, column).value
    #             value_list.append(value)
    #         values.append(value_list)
    #     return values
    #
    # def get_all_value_2(self):
    #     """获取指定表单的所有数据(除去表头)"""
    #     rows_obj = self.work_sheet.iter_rows(min_row=2, max_row=self.work_sheet.max_row,
    #                                          values_only=True)  # 指定values_only 会直接提取数据不需要再使用cell().value
    #     values = []
    #     for row_tuple in rows_obj:
    #         value_list = []
    #         for value in row_tuple:
    #             value_list.append(value)
    #         values.append(value_list)
    #     return values

    def get_excel_title(self):
        """获取sheet表头"""
        title_key = tuple(self.work_sheet.iter_rows(max_row=1, values_only=True))[0]
        return title_key

    # def get_listdict_all_value(self):
    #     """获取所有数据，返回嵌套字典的列表"""
    #     sheet_title = self.get_excel_title()
    #     all_values = self.get_all_value_2()
    #     value_list = []
    #     for value in all_values:
    #         value_list.append(dict(zip(sheet_title, value)))
    #     return value_list
    #
    # def get_list_nametuple_all_value(self):
    #     """获取所有数据，返回嵌套命名元组的列表"""
    #     sheet_title = self.get_excel_title()
    #     values = self.get_all_value_2()
    #
    #     excel = namedtuple('excel', sheet_title)
    #     value_list = []
    #     for value in values:
    #         e = excel(*value)
    #         value_list.append(e)
    #     return value_list


    #获取sheet页所有数据
    # def get_data(self,filename,sheetname):

    def get_all_data(self):
        wb =load_workbook(self.filename)
        # mode = eval(ReadConfig.get_config(case_config_path,"MODE","mode"))
        sheet = wb[self.sheet_name]  # 表单名
        test_data =[]
        for case_id in range(2, sheet.max_row+1):
            row_data = {}  # 字典
            is_run = sheet.cell(case_id, int(ReadConfig.get_test_data("run"))).value
            #通过 excel里的run列判断是否把数据加进DDT去执行
            if is_run == 'yes':
                row_data["case_id"] = sheet.cell(case_id, int(ReadConfig.get_test_data("case_id"))).value
                row_data["module"] = sheet.cell(case_id,int( ReadConfig.get_test_data("module"))).value
                row_data["title"] = sheet.cell(case_id, int(ReadConfig.get_test_data("title"))).value
                # row_data["url"] = sheet.cell(case_id, int(ReadConfig.get_test_data("url"))).value
                #正则替换
                row_data["url"] = DoRegx.do_regx(
                    sheet.cell(case_id, int(ReadConfig.get_test_data("url"))).value)

                # row_data["data"] = sheet.cell(case_id,int(ReadConfig.get_test_data("request_data"))).value
                #非正则字串符替换
                #替换管理员账户
                # if sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value.find("${user}") != -1 and sheet.cell(case_id, int(ReadConfig.get_test_data("request_data"))).value.find("${passwd}") != -1:  # if h后面非空  非零  成立的表达式  都为True，只要是True，if下面的代码都会执行
                #     row_data["data"] = sheet.cell(case_id, int(ReadConfig.get_test_data("request_data"))).value.replace("${user}", getattr(GetData,"admin_user")).replace("${passwd}", str(getattr(GetData,"admin_passwd")))
                # else:
                #     row_data["data"] = sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value

                #正则字符串替换   一行替代多行
                row_data["data"] = DoRegx.do_regx(sheet.cell(case_id, int(ReadConfig.get_test_data("request_data"))).value)
                row_data["run"] = sheet.cell(case_id, int(ReadConfig.get_test_data("run"))).value
                row_data["http_method"] = sheet.cell(case_id,int(ReadConfig.get_test_data("http_method"))).value

                # sql 语句中的字符串替换
                # if sheet.cell(case_id+1, int(ReadConfig.get_test_data("check_sql"))).value !=None:
                #     if sheet.cell(case_id+1,int(ReadConfig.get_test_data("check_sql"))).value.find("${normal_tel}") != -1:
                #         row_data["data"] = sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value.replace("${user}", getattr(GetData, "admin_user"))
                # else:
                #     row_data["data"] = sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value

                row_data["check_sql"] = sheet.cell(case_id, int(ReadConfig.get_test_data("check_sql"))).value
                row_data["expect"] = sheet.cell(case_id, int(ReadConfig.get_test_data("expect"))).value
                row_data["sheet_name"] = self.sheet_name
                test_data.append(row_data)

            # else:
            #     row_data = {}  # 字典
            #     row_data["case_id"] = sheet.cell(case_id+1, 1).value
            #     row_data["username"] = sheet.cell(case_id+1, 2).value
            #     row_data["password"] = sheet.cell(case_id+1, 3).value
            #     row_data["email"] = sheet.cell(i, 4).value
            #     row_data["sheet_name"] = sheetname
            #     test_data.append(row_data)

        return test_data


    # def write_back(filename,sheetname,row,col,result):#回写数据
    def write_back(self,row, col, result):
        try:
            if isinstance(row,int) and isinstance(col,int):
                wb = load_workbook(self.filename)
                sheet = wb[self.sheet_name]
                sheet.cell(row,col).value = result
                wb.save(self.filename)#保存结果
            else:
                raise TypeError("row and col 必须是 int类型")
        except Exception as e:
            raise e



if __name__ == '__main__':
    from dataconfig.project_path import *

    print(DoExcel(test_case_data_path,"login").get_all_data())
